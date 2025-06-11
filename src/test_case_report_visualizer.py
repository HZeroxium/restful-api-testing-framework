"""
Test Case Report visualizer for test case generation results.
Exports JSON test case results to Excel format for easy analysis.
"""

import os
import argparse
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from utils.test_case_report_utils import (
    load_test_case_report,
    analyze_test_case_report,
    generate_constraint_script_mapping,
    generate_test_case_insights,
)
from common.logger import LoggerFactory, LoggerType, LogLevel


class TestCaseReportVisualizer:
    """Visualizes test case generation results and exports to Excel."""

    def __init__(self, verbose: bool = False):
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel functionality requires pandas and openpyxl")

        # Initialize logger
        log_level = LogLevel.DEBUG if verbose else LogLevel.INFO
        self.logger = LoggerFactory.get_logger(
            name="test-case-report-visualizer",
            logger_type=LoggerType.STANDARD,
            level=log_level,
        )

        if not EXCEL_AVAILABLE:
            self.logger.warning(
                "Excel export requires pandas and openpyxl. Install with: pip install pandas openpyxl"
            )

        # Define color schemes
        self.colors = {
            "header": "2E7D32",
            "error": "D32F2F",
            "warning": "F57C00",
            "info": "1976D2",
            "success": "388E3C",
            "light_gray": "F5F5F5",
            "medium_gray": "E0E0E0",
            "constraint_mapped": "E8F5E8",
            "constraint_unmapped": "FFEBEE",
            "test_case": "E3F2FD",
        }

        # Define fonts
        self.fonts = {
            "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "title": Font(name="Calibri", size=14, bold=True),
            "normal": Font(name="Calibri", size=10),
            "bold": Font(name="Calibri", size=10, bold=True),
            "small": Font(name="Calibri", size=9),
        }

    def export_to_excel(
        self,
        test_case_data: Dict[str, Any],
        output_path: str,
        include_detailed_analysis: bool = True,
    ) -> str:
        """
        Export test case results to Excel format.

        Args:
            test_case_data: Test case report data
            output_path: Path for the output Excel file
            include_detailed_analysis: Whether to include detailed analysis sheets

        Returns:
            Path to the created Excel file
        """
        self.logger.info("Starting Excel export of test case report")
        self.logger.add_context(
            output_path=output_path,
            include_detailed_analysis=include_detailed_analysis,
            total_test_cases=test_case_data.get("summary", {}).get(
                "total_test_cases", 0
            ),
        )

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        self.logger.debug("Creating Excel worksheets...")

        # Create sheets
        self._create_overview_sheet(wb, test_case_data)
        self._create_constraint_script_mapping_sheet(wb, test_case_data)
        self._create_test_cases_detail_sheet(wb, test_case_data)
        self._create_constraints_detail_sheet(wb, test_case_data)

        if include_detailed_analysis:
            self._create_validation_scripts_sheet(wb, test_case_data)
            self._create_analysis_sheet(wb, test_case_data)
            self._create_insights_sheet(wb, test_case_data)

        # Save workbook
        wb.save(output_path)

        self.logger.info(f"Excel report successfully created: {output_path}")
        return output_path

    def _create_overview_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create overview summary sheet."""
        ws = wb.create_sheet("Overview", 0)

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "Test Case Generation Report Overview"
        ws["A1"].font = self.fonts["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        # Basic info
        row = 3
        endpoint_info = data.get("endpoint", {})
        endpoint = (
            f"{endpoint_info.get('method', '').upper()} {endpoint_info.get('path', '')}"
        )

        basic_info = [
            ("API Name", data.get("api_name", "")),
            ("Endpoint", endpoint),
            ("Description", endpoint_info.get("description", "")),
            ("Auth Required", "Yes" if endpoint_info.get("auth_required") else "No"),
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
        ]

        for label, value in basic_info:
            if label:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = self.fonts["bold"]
                ws[f"B{row}"] = value
            row += 1

        # Summary statistics
        summary = data.get("summary", {})
        summary_info = [
            ("Summary Statistics", ""),
            ("Total Test Cases", summary.get("total_test_cases", 0)),
            ("Total Constraints", summary.get("total_constraints", 0)),
            ("Total Validation Scripts", summary.get("total_validation_scripts", 0)),
            (
                "Include Invalid Data",
                "Yes" if summary.get("include_invalid_data") else "No",
            ),
            ("", ""),
        ]

        for label, value in summary_info:
            if label:
                if label == "Summary Statistics":
                    ws[f"A{row}"] = label
                    ws[f"A{row}"].font = self.fonts["bold"]
                else:
                    ws[f"A{row}"] = label
                    ws[f"A{row}"].font = self.fonts["normal"]
                    ws[f"B{row}"] = value
            row += 1

        # Constraint breakdown
        constraint_breakdown = summary.get("constraint_breakdown", {})
        breakdown_info = [
            ("Constraint Breakdown", ""),
            (
                "Request Parameters",
                constraint_breakdown.get("request_param_constraints", 0),
            ),
            ("Request Body", constraint_breakdown.get("request_body_constraints", 0)),
            (
                "Response Properties",
                constraint_breakdown.get("response_property_constraints", 0),
            ),
            (
                "Request-Response Correlations",
                constraint_breakdown.get("request_response_constraints", 0),
            ),
        ]

        for label, value in breakdown_info:
            if label == "Constraint Breakdown":
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = self.fonts["bold"]
            else:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = value
            row += 1

        # Analysis summary
        analysis = analyze_test_case_report(data)
        row += 1

        analysis_info = [
            ("Quality Metrics", ""),
            (
                "Constraint Coverage Rate",
                f"{analysis['constraint_coverage']['coverage_rate']}%",
            ),
            (
                "Script Generation Rate",
                f"{analysis['quality_metrics']['script_generation_rate']}%",
            ),
            (
                "Overall Quality Score",
                f"{analysis['quality_metrics']['overall_quality_score']}/100 ({analysis['quality_metrics']['quality_grade']})",
            ),
            ("", ""),
            ("Script Analysis", ""),
            (
                "Unique Scripts",
                analysis["validation_script_analysis"]["unique_scripts"],
            ),
            (
                "Script Duplication Rate",
                f"{analysis['validation_script_analysis']['duplication_rate']}%",
            ),
            (
                "Average Script Complexity",
                analysis["script_complexity"]["average_complexity"],
            ),
        ]

        for label, value in analysis_info:
            if label:
                if label in ["Quality Metrics", "Script Analysis"]:
                    ws[f"A{row}"] = label
                    ws[f"A{row}"].font = self.fonts["bold"]
                else:
                    ws[f"A{row}"] = label
                    ws[f"B{row}"] = value

                    # Color code quality metrics
                    if label == "Overall Quality Score":
                        score = analysis["quality_metrics"]["overall_quality_score"]
                        if score >= 80:
                            ws[f"B{row}"].fill = PatternFill(
                                start_color=self.colors["success"],
                                end_color=self.colors["success"],
                                fill_type="solid",
                            )
                            ws[f"B{row}"].font = Font(color="FFFFFF", bold=True)
                        elif score >= 60:
                            ws[f"B{row}"].fill = PatternFill(
                                start_color=self.colors["warning"],
                                end_color=self.colors["warning"],
                                fill_type="solid",
                            )
                            ws[f"B{row}"].font = Font(color="FFFFFF", bold=True)
                        else:
                            ws[f"B{row}"].fill = PatternFill(
                                start_color=self.colors["error"],
                                end_color=self.colors["error"],
                                fill_type="solid",
                            )
                            ws[f"B{row}"].font = Font(color="FFFFFF", bold=True)
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    def _create_constraint_script_mapping_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Create detailed constraint-to-script mapping sheet."""
        ws = wb.create_sheet("Constraint-Script Mapping")

        # Generate mapping data
        mapping_data = generate_constraint_script_mapping(data)

        if not mapping_data:
            ws["A1"] = "No constraint-script mappings found"
            return

        # Create DataFrame
        df = pd.DataFrame(mapping_data)

        # Add headers
        headers = [
            "Test Case Name",
            "Script Name",
            "Script Type",
            "Constraint Description",
            "Constraint Type",
            "Constraint Severity",
            "Has Mapping",
            "Script Length",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, mapping in enumerate(mapping_data, 2):
            # Test Case Name
            ws.cell(row=row_idx, column=1, value=mapping["test_case_name"])

            # Script Name
            ws.cell(row=row_idx, column=2, value=mapping["script_name"])

            # Script Type
            script_type_cell = ws.cell(
                row=row_idx, column=3, value=mapping["script_type"]
            )
            self._color_code_script_type(script_type_cell, mapping["script_type"])

            # Constraint Description
            ws.cell(row=row_idx, column=4, value=mapping["constraint_description"])

            # Constraint Type
            ws.cell(row=row_idx, column=5, value=mapping["constraint_type"])

            # Constraint Severity
            severity_cell = ws.cell(
                row=row_idx, column=6, value=mapping["constraint_severity"]
            )
            self._color_code_severity(severity_cell, mapping["constraint_severity"])

            # Has Mapping
            mapping_cell = ws.cell(
                row=row_idx,
                column=7,
                value="Yes" if mapping["has_constraint_mapping"] else "No",
            )
            if mapping["has_constraint_mapping"]:
                mapping_cell.fill = PatternFill(
                    start_color=self.colors["constraint_mapped"],
                    end_color=self.colors["constraint_mapped"],
                    fill_type="solid",
                )
            else:
                mapping_cell.fill = PatternFill(
                    start_color=self.colors["constraint_unmapped"],
                    end_color=self.colors["constraint_unmapped"],
                    fill_type="solid",
                )

            # Script Length
            ws.cell(row=row_idx, column=8, value=mapping["validation_code_length"])

        # Auto-adjust column widths
        column_widths = [30, 40, 20, 60, 25, 15, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

        # Add borders
        self._add_borders_to_range(ws, 1, len(mapping_data) + 1, 1, len(headers))

    def _create_test_cases_detail_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Create detailed test cases sheet."""
        ws = wb.create_sheet("Test Cases Detail")

        test_cases = data.get("test_cases", [])
        if not test_cases:
            ws["A1"] = "No test cases found"
            return

        # Headers
        headers = [
            "Test Case ID",
            "Test Case Name",
            "Description",
            "Request Params",
            "Request Body",
            "Expected Status",
            "Validation Scripts Count",
            "Script Types",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )

        # Add test case data
        for row_idx, test_case in enumerate(test_cases, 2):
            ws.cell(row=row_idx, column=1, value=test_case.get("id", ""))
            ws.cell(row=row_idx, column=2, value=test_case.get("name", ""))
            ws.cell(row=row_idx, column=3, value=test_case.get("description", ""))

            # Request params (summarized)
            params = test_case.get("request_params", {})
            params_summary = (
                ", ".join(f"{k}={v}" for k, v in params.items()) if params else "None"
            )
            ws.cell(
                row=row_idx,
                column=4,
                value=(
                    params_summary[:100] + "..."
                    if len(params_summary) > 100
                    else params_summary
                ),
            )

            # Request body
            body = test_case.get("request_body")
            body_summary = "Present" if body else "None"
            ws.cell(row=row_idx, column=5, value=body_summary)

            # Expected status
            ws.cell(
                row=row_idx, column=6, value=test_case.get("expected_status_code", "")
            )

            # Validation scripts count
            scripts = test_case.get("validation_scripts", [])
            ws.cell(row=row_idx, column=7, value=len(scripts))

            # Script types
            script_types = list(
                set(script.get("script_type", "") for script in scripts)
            )
            ws.cell(row=row_idx, column=8, value=", ".join(script_types))

        # Auto-adjust column widths
        column_widths = [25, 40, 50, 40, 15, 15, 20, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

        # Add borders
        self._add_borders_to_range(ws, 1, len(test_cases) + 1, 1, len(headers))

    def _create_constraints_detail_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Create detailed constraints sheet."""
        ws = wb.create_sheet("Constraints Detail")

        constraints = data.get("constraints", [])
        if not constraints:
            ws["A1"] = "No constraints found"
            return

        # Headers
        headers = [
            "Constraint ID",
            "Type",
            "Description",
            "Severity",
            "Source",
            "Constraint Type",
            "Validation Rule",
            "Has Scripts",
            "Script Count",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )

        # Get constraint-to-script mapping for script count
        mapping = generate_constraint_script_mapping(data)
        constraint_script_count = {}
        for item in mapping:
            constraint_id = item["constraint_id"]
            if constraint_id:
                constraint_script_count[constraint_id] = (
                    constraint_script_count.get(constraint_id, 0) + 1
                )

        # Add constraint data
        for row_idx, constraint in enumerate(constraints, 2):
            constraint_id = constraint.get("id", "")

            ws.cell(row=row_idx, column=1, value=constraint_id)
            ws.cell(row=row_idx, column=2, value=constraint.get("type", ""))
            ws.cell(row=row_idx, column=3, value=constraint.get("description", ""))

            # Severity with color coding
            severity_cell = ws.cell(
                row=row_idx, column=4, value=constraint.get("severity", "")
            )
            self._color_code_severity(severity_cell, constraint.get("severity", ""))

            ws.cell(row=row_idx, column=5, value=constraint.get("source", ""))

            details = constraint.get("details", {})
            ws.cell(row=row_idx, column=6, value=details.get("constraint_type", ""))
            ws.cell(row=row_idx, column=7, value=details.get("validation_rule", ""))

            # Has scripts
            script_count = constraint_script_count.get(constraint_id, 0)
            has_scripts_cell = ws.cell(
                row=row_idx, column=8, value="Yes" if script_count > 0 else "No"
            )
            if script_count > 0:
                has_scripts_cell.fill = PatternFill(
                    start_color=self.colors["constraint_mapped"],
                    end_color=self.colors["constraint_mapped"],
                    fill_type="solid",
                )
            else:
                has_scripts_cell.fill = PatternFill(
                    start_color=self.colors["constraint_unmapped"],
                    end_color=self.colors["constraint_unmapped"],
                    fill_type="solid",
                )

            # Script count
            ws.cell(row=row_idx, column=9, value=script_count)

        # Auto-adjust column widths
        column_widths = [25, 20, 60, 15, 15, 20, 25, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

        # Add borders
        self._add_borders_to_range(ws, 1, len(constraints) + 1, 1, len(headers))

    def _create_validation_scripts_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Create detailed validation scripts sheet."""
        ws = wb.create_sheet("Validation Scripts")

        # Collect all scripts
        all_scripts = []
        test_cases = data.get("test_cases", [])

        for test_case in test_cases:
            test_case_name = test_case.get("name", "")
            scripts = test_case.get("validation_scripts", [])
            for script in scripts:
                script_info = {
                    "test_case_name": test_case_name,
                    "script_id": script.get("id", ""),
                    "script_name": script.get("name", ""),
                    "script_type": script.get("script_type", ""),
                    "description": script.get("description", ""),
                    "constraint_id": script.get("constraint_id", ""),
                    "code_length": len(script.get("validation_code", "")),
                    "validation_code": script.get("validation_code", ""),
                }
                all_scripts.append(script_info)

        if not all_scripts:
            ws["A1"] = "No validation scripts found"
            return

        # Headers
        headers = [
            "Test Case",
            "Script ID",
            "Script Name",
            "Script Type",
            "Description",
            "Constraint ID",
            "Code Length",
            "Validation Code",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )

        # Add script data
        for row_idx, script in enumerate(all_scripts, 2):
            ws.cell(row=row_idx, column=1, value=script["test_case_name"])
            ws.cell(row=row_idx, column=2, value=script["script_id"])
            ws.cell(row=row_idx, column=3, value=script["script_name"])

            # Script type with color coding
            script_type_cell = ws.cell(
                row=row_idx, column=4, value=script["script_type"]
            )
            self._color_code_script_type(script_type_cell, script["script_type"])

            ws.cell(row=row_idx, column=5, value=script["description"])
            ws.cell(row=row_idx, column=6, value=script["constraint_id"])
            ws.cell(row=row_idx, column=7, value=script["code_length"])

            # Validation code (truncated for display)
            code = script["validation_code"]
            # truncated_code = code[:200] + "..." if len(code) > 200 else code
            truncated_code = code
            code_cell = ws.cell(row=row_idx, column=8, value=truncated_code)
            code_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-adjust column widths
        column_widths = [25, 25, 40, 20, 40, 25, 15, 50]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

        # Add borders
        self._add_borders_to_range(ws, 1, len(all_scripts) + 1, 1, len(headers))

    def _create_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create analysis sheet with detailed statistics."""
        ws = wb.create_sheet("Analysis")

        analysis = analyze_test_case_report(data)

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "Test Case Generation Analysis"
        ws["A1"].font = self.fonts["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3

        # Quality Metrics Section
        ws[f"A{row}"] = "Quality Metrics"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        quality_metrics = analysis["quality_metrics"]
        quality_data = [
            ("Metric", "Value", "Grade"),
            (
                "Script Generation Rate",
                f"{quality_metrics['script_generation_rate']}%",
                "",
            ),
            (
                "Constraint Coverage Rate",
                f"{quality_metrics['constraint_coverage_rate']}%",
                "",
            ),
            (
                "Test Data Completeness",
                f"{quality_metrics['test_data_completeness']}%",
                "",
            ),
            (
                "Overall Quality Score",
                f"{quality_metrics['overall_quality_score']}/100",
                quality_metrics["quality_grade"],
            ),
        ]

        for i, (metric, value, grade) in enumerate(quality_data):
            if i == 0:  # Header row
                for col_idx, header in enumerate([metric, value, grade], 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = self.fonts["header"]
                    cell.fill = PatternFill(
                        start_color=self.colors["header"],
                        end_color=self.colors["header"],
                        fill_type="solid",
                    )
            else:
                ws[f"A{row}"] = metric
                ws[f"B{row}"] = value
                ws[f"C{row}"] = grade
            row += 1

        row += 2

        # Constraint Analysis Section
        ws[f"A{row}"] = "Constraint Analysis"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        constraint_analysis = analysis["constraint_analysis"]
        constraint_data = [
            ("Constraint Type", "Count", "Percentage"),
        ]

        for constraint_type, count in constraint_analysis["by_type"].items():
            percentage = constraint_analysis["type_distribution"].get(
                constraint_type, 0
            )
            constraint_data.append(
                (constraint_type.replace("_", " ").title(), count, f"{percentage}%")
            )

        for i, (ctype, count, percentage) in enumerate(constraint_data):
            if i == 0:  # Header row
                for col_idx, header in enumerate([ctype, count, percentage], 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = self.fonts["header"]
                    cell.fill = PatternFill(
                        start_color=self.colors["header"],
                        end_color=self.colors["header"],
                        fill_type="solid",
                    )
            else:
                ws[f"A{row}"] = ctype
                ws[f"B{row}"] = count
                ws[f"C{row}"] = percentage
            row += 1

        row += 2

        # Script Complexity Analysis
        ws[f"A{row}"] = "Script Complexity Analysis"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        complexity = analysis["script_complexity"]
        complexity_data = [
            ("Complexity Level", "Count", "Percentage"),
            (
                "Simple Scripts",
                complexity["simple_scripts"],
                f"{complexity['complexity_distribution']['simple']}%",
            ),
            (
                "Medium Scripts",
                complexity["medium_scripts"],
                f"{complexity['complexity_distribution']['medium']}%",
            ),
            (
                "Complex Scripts",
                complexity["complex_scripts"],
                f"{complexity['complexity_distribution']['complex']}%",
            ),
        ]

        for i, (level, count, percentage) in enumerate(complexity_data):
            if i == 0:  # Header row
                for col_idx, header in enumerate([level, count, percentage], 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = self.fonts["header"]
                    cell.fill = PatternFill(
                        start_color=self.colors["header"],
                        end_color=self.colors["header"],
                        fill_type="solid",
                    )
            else:
                ws[f"A{row}"] = level
                ws[f"B{row}"] = count
                ws[f"C{row}"] = percentage
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

    def _create_insights_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create insights sheet with recommendations."""
        ws = wb.create_sheet("Insights")

        # Title
        ws.merge_cells("A1:C1")
        ws["A1"] = "Test Case Generation Insights & Recommendations"
        ws["A1"].font = self.fonts["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3

        # Generate insights
        insights = generate_test_case_insights(data)

        ws[f"A{row}"] = "Key Insights"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        for insight in insights:
            ws[f"A{row}"] = insight
            ws[f"A{row}"].font = self.fonts["normal"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        row += 2

        # Recommendations
        ws[f"A{row}"] = "Recommendations"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        recommendations = self._generate_recommendations(data)
        for recommendation in recommendations:
            ws[f"A{row}"] = recommendation
            ws[f"A{row}"].font = self.fonts["normal"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        # Auto-adjust column width
        ws.column_dimensions["A"].width = 100

    def _color_code_severity(self, cell, severity: str) -> None:
        """Apply color coding based on severity."""
        if severity == "error":
            cell.fill = PatternFill(
                start_color=self.colors["error"],
                end_color=self.colors["error"],
                fill_type="solid",
            )
            cell.font = Font(color="FFFFFF", bold=True)
        elif severity == "warning":
            cell.fill = PatternFill(
                start_color=self.colors["warning"],
                end_color=self.colors["warning"],
                fill_type="solid",
            )
            cell.font = Font(color="FFFFFF", bold=True)
        elif severity == "info":
            cell.fill = PatternFill(
                start_color=self.colors["info"],
                end_color=self.colors["info"],
                fill_type="solid",
            )
            cell.font = Font(color="FFFFFF", bold=True)

    def _color_code_script_type(self, cell, script_type: str) -> None:
        """Apply color coding based on script type."""
        type_colors = {
            "request_param": "E8F5E8",
            "request_body": "FFF3E0",
            "response_property": "E3F2FD",
            "request_response": "F3E5F5",
        }

        color = type_colors.get(script_type, self.colors["light_gray"])
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _add_borders_to_range(
        self, ws, min_row: int, max_row: int, min_col: int, max_col: int
    ) -> None:
        """Add borders to a range of cells."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.border = thin_border

    def _generate_recommendations(self, data: Dict[str, Any]) -> List[str]:
        """Generate actionable recommendations based on analysis."""
        recommendations = []
        analysis = analyze_test_case_report(data)

        self.logger.debug("Generating recommendations based on analysis")

        # Coverage recommendations
        coverage_rate = analysis["constraint_coverage"]["coverage_rate"]
        if coverage_rate < 70:
            recommendation = f"🎯 Low constraint coverage ({coverage_rate}%). Review constraint-to-script mapping and ensure all constraints have corresponding validation scripts"
            recommendations.append(recommendation)
            self.logger.debug(f"Added coverage recommendation: {coverage_rate}%")

        # Quality recommendations
        quality_score = analysis["quality_metrics"]["overall_quality_score"]
        if quality_score < 70:
            recommendation = f"📈 Quality score is {quality_score}/100. Focus on improving script generation rate and constraint coverage"
            recommendations.append(recommendation)
            self.logger.debug(f"Added quality recommendation: {quality_score}/100")

        # Script complexity recommendations
        complexity = analysis["script_complexity"]
        if complexity["complex_scripts"] > complexity["simple_scripts"]:
            recommendation = "🔧 High number of complex scripts detected. Consider refactoring for better maintainability"
            recommendations.append(recommendation)
            self.logger.debug("Added script complexity recommendation")

        # Script duplication recommendations
        duplication_rate = analysis["validation_script_analysis"]["duplication_rate"]
        if duplication_rate > 30:
            recommendation = f"🔄 High script duplication rate ({duplication_rate}%). Consider creating reusable validation functions"
            recommendations.append(recommendation)
            self.logger.debug(f"Added duplication recommendation: {duplication_rate}%")

        # Test data recommendations
        test_data = analysis["test_data_analysis"]
        if test_data["parameter_usage_rate"] < 50:
            recommendation = "📝 Low parameter usage in test data. Consider adding more diverse test scenarios with parameters"
            recommendations.append(recommendation)
            self.logger.debug("Added parameter usage recommendation")

        # Constraint type balance
        constraint_analysis = analysis["constraint_analysis"]
        if constraint_analysis["by_type"].get("request_param", 0) == 0:
            recommendation = "📋 No request parameter constraints found. Consider adding parameter validation rules"
            recommendations.append(recommendation)
            self.logger.debug("Added request parameter constraint recommendation")

        if constraint_analysis["by_type"].get("response_property", 0) == 0:
            recommendation = "📊 No response property constraints found. Consider adding response validation rules"
            recommendations.append(recommendation)
            self.logger.debug("Added response property constraint recommendation")

        # Positive recommendations
        if coverage_rate >= 90 and quality_score >= 80:
            recommendation = "✅ Excellent test case generation! Your API has comprehensive validation coverage"
            recommendations.append(recommendation)
            self.logger.debug("Added positive recommendation for excellent coverage")

        if not recommendations:
            recommendation = "✅ Test case generation completed successfully with good overall quality"
            recommendations.append(recommendation)
            self.logger.debug("Added default positive recommendation")

        self.logger.info(f"Generated {len(recommendations)} recommendations")
        return recommendations


def export_test_case_report_to_excel(
    json_file_path: str,
    output_path: Optional[str] = None,
    include_detailed_analysis: bool = True,
    verbose: bool = False,
) -> str:
    """
    Export test case JSON report to Excel format.

    Args:
        json_file_path: Path to the JSON test case report
        output_path: Optional output path (auto-generated if not provided)
        include_detailed_analysis: Whether to include detailed analysis sheets
        verbose: Enable verbose logging

    Returns:
        Path to the created Excel file
    """
    # Initialize logger
    log_level = LogLevel.DEBUG if verbose else LogLevel.INFO
    logger = LoggerFactory.get_logger(
        name="test-case-excel-exporter",
        logger_type=LoggerType.STANDARD,
        level=log_level,
    )

    if not EXCEL_AVAILABLE:
        error_msg = "Excel export requires pandas and openpyxl. Install with: pip install pandas openpyxl"
        logger.error(error_msg)
        raise ImportError(error_msg)

    logger.info("Starting test case report Excel export")
    logger.add_context(
        input_file=json_file_path,
        output_path=output_path,
        include_detailed_analysis=include_detailed_analysis,
    )

    # Load the JSON data
    test_case_data = load_test_case_report(json_file_path)

    if "error" in test_case_data:
        error_msg = f"Error loading test case report: {test_case_data['error']}"
        logger.error(error_msg)
        raise ValueError(error_msg)

    # Generate output path if not provided
    if output_path is None:
        base_name = os.path.splitext(os.path.basename(json_file_path))[0]
        output_dir = os.path.dirname(json_file_path)
        output_path = os.path.join(output_dir, f"{base_name}_report.xlsx")
        logger.debug(f"Auto-generated output path: {output_path}")

    # Create visualizer and export
    visualizer = TestCaseReportVisualizer(verbose=verbose)
    result_path = visualizer.export_to_excel(
        test_case_data, output_path, include_detailed_analysis
    )

    logger.info("Excel export completed successfully")
    logger.add_context(result_path=result_path)

    return result_path


def main():
    """Main function for command-line usage."""
    parser = argparse.ArgumentParser(
        description="Export test case JSON results to Excel"
    )
    parser.add_argument("input_file", help="Path to the test case JSON file")
    parser.add_argument(
        "--output",
        "-o",
        help="Output Excel file path (auto-generated if not specified)",
    )
    parser.add_argument(
        "--no-analysis", action="store_true", help="Skip detailed analysis sheets"
    )
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")

    args = parser.parse_args()

    # Initialize logger
    log_level = LogLevel.DEBUG if args.verbose else LogLevel.INFO
    logger = LoggerFactory.get_logger(
        name="test-case-report-visualizer-main",
        logger_type=LoggerType.STANDARD,
        level=log_level,
    )

    if not EXCEL_AVAILABLE:
        logger.error("Excel export requires pandas and openpyxl")
        logger.error("Install with: pip install pandas openpyxl")
        return 1

    logger.info("Starting test case report visualizer")
    logger.add_context(
        input_file=args.input_file,
        output_file=args.output,
        skip_analysis=args.no_analysis,
        verbose=args.verbose,
    )

    try:
        output_path = export_test_case_report_to_excel(
            args.input_file,
            args.output,
            include_detailed_analysis=not args.no_analysis,
            verbose=args.verbose,
        )

        logger.info("Excel report created successfully")
        logger.add_context(input_file=args.input_file, output_file=output_path)

        # Final status using logger
        logger.info(f"✅ Excel report created successfully:")
        logger.info(f"   Input:  {args.input_file}")
        logger.info(f"   Output: {output_path}")

        if args.verbose:
            # Load and show summary
            data = load_test_case_report(args.input_file)
            summary = data.get("summary", {})
            endpoint = data.get("endpoint", {})
            endpoint_name = (
                f"{endpoint.get('method', '').upper()} {endpoint.get('path', '')}"
            )

            logger.info("📊 Report Summary:")
            logger.info(f"   Endpoint: {endpoint_name}")
            logger.info(f"   Test Cases: {summary.get('total_test_cases', 0)}")
            logger.info(f"   Constraints: {summary.get('total_constraints', 0)}")
            logger.info(
                f"   Validation Scripts: {summary.get('total_validation_scripts', 0)}"
            )

        return 0

    except Exception as e:
        logger.error(f"Error creating Excel report: {str(e)}")
        if args.verbose:
            import traceback

            logger.debug(f"Traceback: {traceback.format_exc()}")
        return 1


if __name__ == "__main__":
    exit(main())
