# src/report_visualizer.py

"""
Report visualizer for constraint mining results.
Exports JSON constraint mining results to Excel format for easy analysis.
"""

import os
import argparse
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from utils.report_utils import (
    load_constraint_mining_report,
    analyze_constraint_mining_result,
    generate_constraint_insights,
)
from common.logger import LoggerFactory, LoggerType, LogLevel


class ConstraintReportVisualizer:
    """Visualizes constraint mining results and exports to Excel."""

    def __init__(self, verbose: bool = False):
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel functionality requires pandas and openpyxl")

        # Initialize logger
        log_level = LogLevel.DEBUG if verbose else LogLevel.INFO
        self.logger = LoggerFactory.get_logger(
            name="constraint-report-visualizer",
            logger_type=LoggerType.STANDARD,
            level=log_level,
        )

        if not EXCEL_AVAILABLE:
            self.logger.warning(
                "Excel export requires pandas and openpyxl. Install with: pip install pandas openpyxl"
            )

        # Define color schemes
        self.colors = {
            "header": "366092",
            "error": "D32F2F",
            "warning": "F57C00",
            "info": "1976D2",
            "success": "388E3C",
            "light_gray": "F5F5F5",
            "medium_gray": "E0E0E0",
        }

        # Define fonts
        self.fonts = {
            "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "title": Font(name="Calibri", size=14, bold=True),
            "normal": Font(name="Calibri", size=10),
            "bold": Font(name="Calibri", size=10, bold=True),
        }

    def export_to_excel(
        self,
        constraint_data: Dict[str, Any],
        output_path: str,
        include_analysis: bool = True,
    ) -> str:
        """
        Export constraint mining results to Excel format.

        Args:
            constraint_data: Constraint mining result data
            output_path: Path for the output Excel file
            include_analysis: Whether to include analysis sheets

        Returns:
            Path to the created Excel file
        """
        self.logger.info("Starting Excel export of constraint mining report")
        self.logger.add_context(
            output_path=output_path,
            include_analysis=include_analysis,
            total_constraints=constraint_data.get("total_constraints", 0),
        )

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        self.logger.debug("Creating Excel worksheets...")

        # Create sheets
        self._create_summary_sheet(wb, constraint_data)
        self._create_constraints_detail_sheet(wb, constraint_data)

        if include_analysis:
            self._create_analysis_sheet(wb, constraint_data)
            self._create_insights_sheet(wb, constraint_data)

        # Save workbook
        wb.save(output_path)

        self.logger.info(f"Excel report successfully created: {output_path}")
        return output_path

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create summary overview sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "Constraint Mining Report Summary"
        ws["A1"].font = self.fonts["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        # Basic info
        row = 3
        endpoint = (
            f"{data.get('endpoint_method', '').upper()} {data.get('endpoint_path', '')}"
        )

        info_data = [
            ("Endpoint", endpoint),
            ("Total Constraints", data.get("total_constraints", 0)),
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Constraint Breakdown", ""),
            ("Request Parameters", len(data.get("request_param_constraints", []))),
            ("Request Body", len(data.get("request_body_constraints", []))),
            ("Response Properties", len(data.get("response_property_constraints", []))),
            (
                "Request-Response Correlations",
                len(data.get("request_response_constraints", [])),
            ),
        ]

        for label, value in info_data:
            if label:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = self.fonts["bold"]
                ws[f"B{row}"] = value
            row += 1

        # Severity distribution
        row += 1
        ws[f"A{row}"] = "Severity Distribution"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 1

        severity_counts = self._count_severities(data)
        for severity, count in severity_counts.items():
            ws[f"A{row}"] = severity.title()
            ws[f"B{row}"] = count

            # Color code severity
            if severity == "error":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["error"],
                    end_color=self.colors["error"],
                    fill_type="solid",
                )
                ws[f"A{row}"].font = Font(color="FFFFFF", bold=True)
            elif severity == "warning":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["warning"],
                    end_color=self.colors["warning"],
                    fill_type="solid",
                )
                ws[f"A{row}"].font = Font(color="FFFFFF", bold=True)
            elif severity == "info":
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.colors["info"],
                    end_color=self.colors["info"],
                    fill_type="solid",
                )
                ws[f"A{row}"].font = Font(color="FFFFFF", bold=True)

            row += 1

        # Mining quality
        if "result" in data:
            row += 1
            ws[f"A{row}"] = "Mining Quality"
            ws[f"A{row}"].font = self.fonts["bold"]
            row += 1

            mining_results = data["result"].get("mining_results", {})
            for miner, result in mining_results.items():
                status = result.get("status", "unknown")
                source = result.get("source", "unknown")
                ws[f"A{row}"] = f"{miner.replace('_', ' ').title()}"
                ws[f"B{row}"] = f"{status} ({source})"

                # Color code status
                if status == "success":
                    ws[f"B{row}"].fill = PatternFill(
                        start_color=self.colors["success"],
                        end_color=self.colors["success"],
                        fill_type="solid",
                    )
                    ws[f"B{row}"].font = Font(color="FFFFFF")
                elif status == "failed":
                    ws[f"B{row}"].fill = PatternFill(
                        start_color=self.colors["error"],
                        end_color=self.colors["error"],
                        fill_type="solid",
                    )
                    ws[f"B{row}"].font = Font(color="FFFFFF")

                row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _create_constraints_detail_sheet(
        self, wb: Workbook, data: Dict[str, Any]
    ) -> None:
        """Create detailed constraints sheet."""
        ws = wb.create_sheet("Constraints Detail")

        # Prepare data for DataFrame
        all_constraints = []

        # Process each constraint category
        categories = [
            ("request_param_constraints", "Request Parameter"),
            ("request_body_constraints", "Request Body"),
            ("response_property_constraints", "Response Property"),
            ("request_response_constraints", "Request-Response"),
        ]

        for category_key, category_name in categories:
            constraints = data.get(category_key, [])
            for constraint in constraints:
                constraint_row = {
                    "Category": category_name,
                    "ID": constraint.get("id", ""),
                    "Description": constraint.get("description", ""),
                    "Severity": constraint.get("severity", ""),
                    "Source": constraint.get("source", ""),
                    "Type": constraint.get("type", ""),
                    "Constraint_Type": constraint.get("details", {}).get(
                        "constraint_type", ""
                    ),
                    "Validation_Rule": constraint.get("details", {}).get(
                        "validation_rule", ""
                    ),
                    "Details": self._format_details(constraint.get("details", {})),
                }
                all_constraints.append(constraint_row)

        if not all_constraints:
            ws["A1"] = "No constraints found"
            return

        # Create DataFrame and add to sheet
        df = pd.DataFrame(all_constraints)

        # Add headers
        headers = [
            "Category",
            "ID",
            "Description",
            "Severity",
            "Source",
            "Type",
            "Constraint Type",
            "Validation Rule",
            "Details",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.font = self.fonts["normal"]

                # Color code severity
                if col_name == "Severity":
                    if value == "error":
                        cell.fill = PatternFill(
                            start_color=self.colors["error"],
                            end_color=self.colors["error"],
                            fill_type="solid",
                        )
                        cell.font = Font(color="FFFFFF")
                    elif value == "warning":
                        cell.fill = PatternFill(
                            start_color=self.colors["warning"],
                            end_color=self.colors["warning"],
                            fill_type="solid",
                        )
                        cell.font = Font(color="FFFFFF")
                    elif value == "info":
                        cell.fill = PatternFill(
                            start_color=self.colors["info"],
                            end_color=self.colors["info"],
                            fill_type="solid",
                        )
                        cell.font = Font(color="FFFFFF")

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                20
            )

        # Special widths for specific columns
        ws.column_dimensions["C"].width = 50  # Description
        ws.column_dimensions["I"].width = 30  # Details

        # Add borders
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=1, max_row=len(all_constraints) + 1, min_col=1, max_col=len(headers)
        ):
            for cell in row:
                cell.border = thin_border

    def _create_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create analysis sheet with statistics."""
        ws = wb.create_sheet("Analysis")

        analysis = analyze_constraint_mining_result(data)

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "Constraint Mining Analysis"
        ws["A1"].font = self.fonts["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3

        # Severity Analysis
        ws[f"A{row}"] = "Severity Analysis"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        severity_data = analysis["severity_analysis"]
        headers = ["Severity", "Count", "Percentage"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )

        row += 1
        for severity, count in severity_data["counts"].items():
            if count > 0:
                percentage = severity_data["percentages"].get(severity, 0)
                ws[f"A{row}"] = severity.title()
                ws[f"B{row}"] = count
                ws[f"C{row}"] = f"{percentage}%"
                row += 1

        row += 2

        # Constraint Types Analysis
        ws[f"A{row}"] = "Top Constraint Types"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        headers = ["Constraint Type", "Count"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.fonts["header"]
            cell.fill = PatternFill(
                start_color=self.colors["header"],
                end_color=self.colors["header"],
                fill_type="solid",
            )

        row += 1
        for constraint_type, count in analysis["constraint_types"]["most_common"]:
            ws[f"A{row}"] = constraint_type.replace("_", " ").title()
            ws[f"B{row}"] = count
            row += 1

        row += 2

        # Mining Quality
        ws[f"A{row}"] = "Mining Quality Metrics"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        quality = analysis["mining_quality"]
        quality_metrics = [
            ("Successful Miners", quality["successful_miners"]),
            ("Failed Miners", quality["failed_miners"]),
            ("Skipped Miners", quality["skipped_miners"]),
            ("LLM Sourced", quality["llm_sourced"]),
            ("Fallback Sourced", quality["fallback_sourced"]),
        ]

        for metric, value in quality_metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_insights_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create insights sheet with recommendations."""
        ws = wb.create_sheet("Insights")

        # Title
        ws.merge_cells("A1:C1")
        ws["A1"] = "Constraint Mining Insights & Recommendations"
        ws["A1"].font = self.fonts["title"]
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3

        # Generate insights
        insights = generate_constraint_insights(data)

        ws[f"A{row}"] = "Key Insights"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        for insight in insights:
            ws[f"A{row}"] = insight
            ws[f"A{row}"].font = self.fonts["normal"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        row += 2

        # Recommendations
        ws[f"A{row}"] = "Recommendations"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        recommendations = self._generate_recommendations(data)
        for recommendation in recommendations:
            ws[f"A{row}"] = recommendation
            ws[f"A{row}"].font = self.fonts["normal"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        # Auto-adjust column width
        ws.column_dimensions["A"].width = 80

    def _count_severities(self, data: Dict[str, Any]) -> Dict[str, int]:
        """Count constraints by severity."""
        severity_counts = {"error": 0, "warning": 0, "info": 0, "unknown": 0}

        categories = [
            "request_param_constraints",
            "request_body_constraints",
            "response_property_constraints",
            "request_response_constraints",
        ]

        for category in categories:
            constraints = data.get(category, [])
            for constraint in constraints:
                severity = constraint.get("severity", "unknown")
                if severity in severity_counts:
                    severity_counts[severity] += 1
                else:
                    severity_counts["unknown"] += 1

        return severity_counts

    def _format_details(self, details: Dict[str, Any]) -> str:
        """Format constraint details as readable string."""
        if not details:
            return ""

        formatted_parts = []
        for key, value in details.items():
            if key in ["constraint_type", "validation_rule"]:
                continue  # These are shown in separate columns

            if isinstance(value, list):
                value = ", ".join(map(str, value))

            formatted_parts.append(f"{key.replace('_', ' ').title()}: {value}")

        return "; ".join(formatted_parts)

    def _generate_recommendations(self, data: Dict[str, Any]) -> List[str]:
        """Generate actionable recommendations."""
        recommendations = []
        analysis = analyze_constraint_mining_result(data)

        self.logger.debug("Generating recommendations based on constraint analysis")

        # Check for missing constraints
        total = analysis["constraint_summary"]["total_constraints"]
        if total == 0:
            recommendations.append(
                "🎯 Consider adding descriptions and examples to your OpenAPI specification to improve constraint detection"
            )
            recommendations.append(
                "📝 Review your API documentation for implicit validation rules that could be made explicit"
            )
            self.logger.debug("Added recommendations for missing constraints")
            return recommendations

        # Error severity recommendations
        error_count = analysis["severity_analysis"]["counts"]["error"]
        if error_count > 5:
            recommendation = f"🔴 High number of error-level constraints ({error_count}). Prioritize implementing validation for these critical requirements"
            recommendations.append(recommendation)
            self.logger.debug(
                f"Added error severity recommendation: {error_count} errors"
            )

        # Missing request validation
        if analysis["constraint_summary"]["request_param_count"] == 0:
            recommendation = "📝 No request parameter constraints found. Consider adding parameter validation rules"
            recommendations.append(recommendation)
            self.logger.debug("Added request parameter validation recommendation")

        # Missing response validation
        if analysis["constraint_summary"]["response_property_count"] == 0:
            recommendation = "📋 No response property constraints found. Consider defining response schema validation"
            recommendations.append(recommendation)
            self.logger.debug("Added response property validation recommendation")

        # Low LLM success rate
        quality = analysis["mining_quality"]
        if quality["fallback_sourced"] > quality["llm_sourced"]:
            recommendation = "🤖 LLM analysis had limited success. Consider improving OpenAPI descriptions and examples"
            recommendations.append(recommendation)
            self.logger.debug("Added LLM success rate recommendation")

        # Good coverage
        if analysis["constraint_types"]["unique_types"] > 7:
            recommendation = "✅ Excellent constraint diversity detected. Your API has comprehensive validation coverage"
            recommendations.append(recommendation)
            self.logger.debug("Added positive recommendation for constraint diversity")

        if not recommendations:
            recommendation = (
                "✅ Constraint mining completed successfully with good coverage"
            )
            recommendations.append(recommendation)
            self.logger.debug("Added default positive recommendation")

        self.logger.info(f"Generated {len(recommendations)} recommendations")
        return recommendations


def export_constraint_report_to_excel(
    json_file_path: str,
    output_path: Optional[str] = None,
    include_analysis: bool = True,
    verbose: bool = False,
) -> str:
    """
    Export constraint mining JSON report to Excel format.

    Args:
        json_file_path: Path to the JSON constraint mining report
        output_path: Optional output path (auto-generated if not provided)
        include_analysis: Whether to include analysis sheets
        verbose: Enable verbose logging

    Returns:
        Path to the created Excel file
    """
    # Initialize logger
    log_level = LogLevel.DEBUG if verbose else LogLevel.INFO
    logger = LoggerFactory.get_logger(
        name="constraint-excel-exporter",
        logger_type=LoggerType.STANDARD,
        level=log_level,
    )

    if not EXCEL_AVAILABLE:
        error_msg = "Excel export requires pandas and openpyxl. Install with: pip install pandas openpyxl"
        logger.error(error_msg)
        raise ImportError(error_msg)

    logger.info("Starting constraint mining report Excel export")
    logger.add_context(
        input_file=json_file_path,
        output_path=output_path,
        include_analysis=include_analysis,
    )

    # Load the JSON data
    constraint_data = load_constraint_mining_report(json_file_path)

    if "error" in constraint_data:
        error_msg = f"Error loading constraint report: {constraint_data['error']}"
        logger.error(error_msg)
        raise ValueError(error_msg)

    # Generate output path if not provided
    if output_path is None:
        base_name = os.path.splitext(os.path.basename(json_file_path))[0]
        output_dir = os.path.dirname(json_file_path)
        output_path = os.path.join(output_dir, f"{base_name}_report.xlsx")
        logger.debug(f"Auto-generated output path: {output_path}")

    # Create visualizer and export
    visualizer = ConstraintReportVisualizer(verbose=verbose)
    result_path = visualizer.export_to_excel(
        constraint_data, output_path, include_analysis
    )

    logger.info("Excel export completed successfully")
    logger.add_context(result_path=result_path)

    return result_path


def main():
    """Main function for command-line usage."""
    parser = argparse.ArgumentParser(
        description="Export constraint mining JSON results to Excel"
    )
    parser.add_argument("input_file", help="Path to the constraint mining JSON file")
    parser.add_argument(
        "--output",
        "-o",
        help="Output Excel file path (auto-generated if not specified)",
    )
    parser.add_argument(
        "--no-analysis", action="store_true", help="Skip analysis sheets"
    )
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")

    args = parser.parse_args()

    # Initialize logger
    log_level = LogLevel.DEBUG if args.verbose else LogLevel.INFO
    logger = LoggerFactory.get_logger(
        name="constraint-report-visualizer-main",
        logger_type=LoggerType.STANDARD,
        level=log_level,
    )

    if not EXCEL_AVAILABLE:
        logger.error("Excel export requires pandas and openpyxl")
        logger.error("Install with: pip install pandas openpyxl")
        return 1

    logger.info("Starting constraint mining report visualizer")
    logger.add_context(
        input_file=args.input_file,
        output_file=args.output,
        skip_analysis=args.no_analysis,
        verbose=args.verbose,
    )

    try:
        output_path = export_constraint_report_to_excel(
            args.input_file,
            args.output,
            include_analysis=not args.no_analysis,
            verbose=args.verbose,
        )

        logger.info("Excel report created successfully")
        logger.add_context(input_file=args.input_file, output_file=output_path)

        # Final status using logger
        logger.info("✅ Excel report created successfully:")
        logger.info(f"   Input:  {args.input_file}")
        logger.info(f"   Output: {output_path}")

        if args.verbose:
            # Load and show summary
            data = load_constraint_mining_report(args.input_file)
            total = data.get("total_constraints", 0)
            endpoint = f"{data.get('endpoint_method', '').upper()} {data.get('endpoint_path', '')}"

            logger.info("📊 Report Summary:")
            logger.info(f"   Endpoint: {endpoint}")
            logger.info(f"   Total Constraints: {total}")

        return 0

    except Exception as e:
        logger.error(f"Error creating Excel report: {str(e)}")
        if args.verbose:
            import traceback

            logger.debug(f"Traceback: {traceback.format_exc()}")
        return 1


if __name__ == "__main__":
    exit(main())
